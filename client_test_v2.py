import concurrent
import configparser
import datetime
from functools import partial
import importlib
import json
import logging
import os
import sys
import threading
import time
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, PatternFill, Side
from openpyxl.styles.fonts import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import RowDimension
import serial.tools.list_ports
from PyQt5 import QtCore
from PyQt5 import QtGui, QtWidgets
from PyQt5 import uic
from PyQt5.QtCore import QObject, Qt, pyqtSignal
from PyQt5.QtGui import QColor, QFont, QIcon, QPainter, QPalette, QStandardItem, QStandardItemModel, QTextCharFormat
from PyQt5.QtWidgets import QApplication, QCheckBox, QFileDialog, QLabel, QMessageBox, QTextEdit, QVBoxLayout
from PyQt5.QtGui import QColor, QTextCharFormat, QTextCursor  # 正确导入QTextCursor
from pymodbus import FramerType
from pymodbus.client import ModbusSerialClient
from pymodbus.exceptions import ConnectionException, ModbusIOException

# 设置日志级别为INFO，获取日志记录器实例
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)

# 检查log文件夹是否存在，如果不存在则创建
log_folder = "./log"
if not os.path.exists(log_folder):
    os.makedirs(log_folder)

# 创建一个文件处理器，用于将日志写入文件
file_handler = logging.FileHandler('./log/ClientTest_log.txt',encoding='utf-8')
file_handler.setLevel(logging.INFO)

# 创建一个日志格式
log_format = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
file_handler.setFormatter(log_format)

# 将文件处理器添加到日志记录器
logger.addHandler(file_handler)

stream_handler = logging.StreamHandler(stream=sys.stdout)
logger.addHandler(stream_handler)

class ClientTest(QtCore.QObject):
    """
    定义客户端类ClientTest，用户操作界面
    """
    client_version = 'V1.0'
    release_date_str = '2024-11-21'
    # 以下定义Treeview表格的列名
    STR_PORT = '端口号'
    STR_DEVICE_NAME = '设备名称'
    STR_SOFTWARE_VERSION = '软件版本'
    STR_DEVICE_ID = '设备ID'
    STR_CONNECT_STATUS = '连接状态'
    STR_TEST_PROGRESS = '测试进展'
    STR_TEST_RESULT = '测试结果'

    # 表头
    HEADS = [
        STR_PORT,
        STR_DEVICE_NAME,
        STR_SOFTWARE_VERSION,
        STR_DEVICE_ID,
        STR_CONNECT_STATUS,
        STR_TEST_PROGRESS,
        STR_TEST_RESULT
    ]

    # # 定义设备连接状态，0设备已连接，1设备未连接
    # DEVICE_CONNECT_STATUS = 0
    # CONNECT_STATUS = {
    #     0: '已连接',
    #     1: '未连接'
    # }

    # # 定义测试结果，0未开始测试，1测试进行中，2测试通过，3测试不通过
    # DEVICE_TEST_RESULT = 0
    # TEST_RESULT = {
    #     0: '未开始',
    #     1: '进行中',
    #     2: '通过',
    #     3: '不通过'
    # }
    # window 相关属性
    window_name = '测试客户端'
    win_position_x = 200
    win_position_y = 100
    current_ui_enable = 'n'
    log_ui_enable = 'n'
    

    # 老化时间选项
    aging_duration_options = ['0.5', '1', '1.5', '3', '8', '12', '24']
    selected_aging_duration = '0.5'
    unit_duration = 5.11 # #aging 5.11,current 24.28,stress 85.11,modbus 805.09
    offset_duration = 0
    update_port_enable = True
    max_port_num = 32
    timeout = 30
    no_used_port = '无可用端口'
    port_names = [no_used_port]
    node_ids = [2]
    select_port_names =[]
    check_box_list = []
    # 假设这是全局的设备信息列表
    devices_info_list = []
    last_refresh_time = 0
    
    # 定义电流界面组件list
    label_com_list = []
    editText_current_list = []
    
    timer_running = False
    running = False
    script_name = None
    overall_result = []
    result = '不通过'
    report_title = '测试报告'
    
    log_enable = 'y'
    
    def custom_logger(self, level='INFO', message=''):
        """
        自定义日志输出函数
        :param message: 要记录的日志消息内容
        :param level: 日志级别，可选值为 'DEBUG'、'INFO'、'WARNING'、'ERROR'、'CRITICAL'，默认是 'INFO'
        """
        level = level.upper()
        
        if self.log_enable == 'n':
            return
            
        if level == 'DEBUG':
            logger.debug(message)
        elif level == 'INFO':
            logger.info(message)
        elif level == 'WARNING':
            logger.warning(message)
        elif level == 'ERROR':
            logger.error(message)
        elif level == 'CRITICAL':
            logger.critical(message)
        else:
            raise ValueError("无效的日志级别，请选择 'DEBUG'、'INFO'、'WARNING'、'ERROR'、'CRITICAL' 之一")
    
    class ConfigReader:
        """
        定义一个读取配置文件config.ini的工具类
        """

        def __init__(self, config_file_path):
            self.config_file_path = config_file_path
            self.config = configparser.ConfigParser()
            self.config.read(self.config_file_path, encoding='UTF-8')

        def get_value(self, section, key):
            """
            获取指定 section 和 key 对应的配置值。

            参数：
            - section：配置文件中的节名。
            - key：节中的键名。

            返回值：
            - 对应键的值，如果找不到则返回 None。
            """
            try:
                return self.config.get(section, key)
            except (configparser.NoSectionError, configparser.NoOptionError):
                return None

        def get_section(self, section):
            """
            获取指定 section 下的所有配置项。

            参数：
            - section：配置文件中的节名。

            返回值：
            - 一个字典，包含该节下的所有键值对，如果找不到该节则返回 None。
            """
            try:
                return dict(self.config.items(section))
            except configparser.NoSectionError:
                return None
            
    class StdoutRedirector(QObject):
        """
        用于重定向标准输出并在指定的QTextEdit控件中显示文本，支持文本颜色设置和光标操作。
        """
        update_text_signal = pyqtSignal(QTextCursor, str)  # 用于发送要显示的文本内容信号

        def __init__(self, text_widget: QTextEdit):
            super().__init__()
            self.text_widget = text_widget
            self.text_widget.setReadOnly(True)
            self.buffer = ""
            self.lock = threading.Lock()

            # 在初始化时连接信号与对应的槽函数（这里是同一个类内的方法）
            self.update_text_signal.connect(self.handle_text_update)

        def write(self, string):
            """
            接收要输出的文本内容，处理后显示到文本框中，并根据内容设置颜色，最后将光标移到末尾。
            """
            try:
                with self.lock:
                    self.buffer = ""  # 先清空缓冲区
                    self.buffer += string
                    lines = self.buffer.split('\n')
                    cursor = self.text_widget.textCursor()
                    for line in lines[: -1]:
                        self._append_formatted_text(cursor, line)
                    self.buffer = lines[-1]
                    cursor.movePosition(QTextCursor.End)
                    self.text_widget.setTextCursor(cursor)

                    # 发射信号，传递当前光标和剩余的缓冲区文本内容
                    self.update_text_signal.emit(cursor, self.buffer)
            except Exception as e:
                logger.error(f"Error in write method: {e}")

        def handle_text_update(self, cursor: QTextCursor, text: str):
            """
            接收信号后根据传递过来的光标和文本内容进行后续操作，这里主要是继续处理剩余文本内容。
            此方法作为update_text_signal信号的槽函数，处理接收到的光标和文本信息。
            """
            if text:  # 如果还有剩余文本内容，则继续处理
                self._append_formatted_text(cursor, text)
                cursor.movePosition(QTextCursor.End)
                self.text_widget.setTextCursor(cursor)

        def _append_formatted_text(self, cursor: QTextCursor, text: str):
            """
            根据文本内容设置颜色格式，并将文本添加到光标所在位置。
            """
            color = self._get_text_color(text)
            color_format = QTextCharFormat()
            color_format.setForeground(QColor(color))
            cursor.insertText(text + '\n', color_format)

        def _get_text_color(self, text: str):
            """
            根据文本内容判断对应的颜色。
            """
            if "不通过" in text:
                return "red"
            elif "通过" in text:
                return "green"
            return "black"
        

    def __init__(self):
        super().__init__()
        self.app = QApplication([])
        self.window = uic.loadUi(uifile="ui/client.ui")
        self.read_configfile()
        self.set_window_style()
        self.create_style()
        self.init_widgets()
        # 连接窗口关闭事件到自定义的关闭处理函数
        self.window.closeEvent = self.close_event_handler
        self.window.show()
        self.app.exec_()
        
    def close_event_handler(self, event):
        self.running = False
        self.write_to_json_file(stop_test=True,pause_test=True)

    def write_to_json_file(self, stop_test, pause_test):
        data_to_write = {
            "stop_test": stop_test,
            "pause_test": pause_test
        }
        with open('shared_data.json', 'w') as f:
            json.dump(data_to_write, f)
            
    def read_from_json_file(self):
        """
        从json文件读取标志位，判断是否继续执行测试
        Returns:
            返回fasle则终止测试，true继续测试
        """
        try:
            with open('shared_data.json', 'r') as f:
                data = json.load(f)
                stop_test = data['stop_test']
                pause_test = data['pause_test']
                return stop_test,pause_test
        except FileNotFoundError:
            logger.error("共享的json文件不存在")
            return False,False
        except json.JSONDecodeError:
            logger.error("json文件数据格式错误")
            return False,False
        except Exception as e:
            logger.error(f"读取JSON文件时出现其他未知错误: {e}")
            return False, False

    def set_window_style(self):
        self.window.setFixedSize(1369, 827)
        self.window.setWindowOpacity(1)
        self.window.move(self.win_position_x, self.win_position_y)
        palette = self.window.palette()
        palette.setColor(QPalette.Window, Qt.lightGray)
        try:
            # 从文件路径加载图标
            current_dir = os.getcwd()
            config_file_name = "icon/logo.png"
            config_file_path = os.path.join(current_dir, config_file_name)
            icon = QIcon(config_file_path)
            self.window.setWindowIcon(icon)
        except Exception as e:
            logger.error(e)
        self.window.setWindowTitle(self.window_name)
        
    def read_configfile(self):
        try:
            current_dir = os.getcwd()
            config_file_name = "config/config.ini"
            config_file_path = os.path.join(current_dir, config_file_name)
            config = self.ConfigReader(config_file_path)

            self.window_name = config.get_value('window_parameter', 'window_name').strip("'")
            # logger.info(f'客户端名称：{self.window_name}')
            
            self.current_ui_enable = config.get_value('window_parameter', 'current_ui_enable')
            self.log_ui_enable = config.get_value('window_parameter', 'log_ui_enable')
            # logger.info(f'current_ui_enable= {self.current_ui_enable},log_ui_enable = {self.log_ui_enable}')

            self.win_position_x = int(config.get_value('window_parameter', 'postion_x'))
            self.win_position_y = int(config.get_value('window_parameter', 'postion_y'))
            # logger.info(f'window默认位置：{self.win_position_x, self.win_position_y}')

            self.max_port_num = int(config.get_value('aging_parameter', 'max_port_num'))
            # logger.info(f'端口数目：{self.max_port_num}')

            options_str = config.get_value('aging_parameter', 'aging_options')
            self.aging_duration_options = [x.strip("'") for x in options_str.strip("'").split(", ")]
            self.selected_aging_duration = self.aging_duration_options[0]
            # logger.info(f'老化时间选项：{self.aging_duration_options}')
            
            self.unit_duration = config.get_value('aging_parameter', 'unit_duration')
            self.offset_duration = self.get_offset_duration()
            
            self.time_out = int(config.get_value('aging_parameter', 'time_out'))
            
        except Exception as e:
            logger.error(e)

    def get_offset_duration(self):
        float_result = (float(self.selected_aging_duration) * 3600) / float(self.unit_duration)
        decimal_part = float_result - int(float_result)
        # 将小数部分从秒转换为小时，因为1小时 = 3600秒，所以除以3600
        offset_duration_in_hours = (1-decimal_part) * float(self.unit_duration) / 3600
        # 使用round函数保留两位小数
        logger.info(f'offset_duration_in_hours= {round(offset_duration_in_hours, 5)}')
        return round(offset_duration_in_hours, 5)

    def create_style(self):
        """用于定义控件样式
        """
        # # 定义样式表字符串，只针对窗口本身设置样式
        # self.window_style_sheet = "QWidget#your_window_object_name {" + \
        #              "background-color: #f5f5f5; /* 设置窗口背景色为淡灰色 */" + \
        #              "color: #333; /* 设置窗口内文本颜色为深灰色 */" + \
        #              "}"

        # 表头样式表
        self.str_header_format = "QHeaderView::section {" + \
                                 "background-color: #e9e9e9; /* 淡灰色背景 */" + \
                                 "color: #333; /* 深灰色文本颜色 */" + \
                                 "border-bottom: 1px solid #ccc; /* 底部有一条浅灰色细线分隔 */" + \
                                 "padding: 4px 8px; /* 内边距，上下 4 像素，左右 8 像素 */" + \
                                 "font-size: 16px; /* 较小的字体尺寸 */" + \
                                 "font-weight: normal; /* 正常字体粗细 */" + \
                                 "}"

        # 数据区域样式表
        self.str_data_format = "QTableView {" + \
                               "background-color: white; /* 数据区域设置为白色背景 */" + \
                               "alternate-background-color: #f5f5f5; /* 隔行变色，设置为淡灰色 */" + \
                               "color: #333; /* 数据文本颜色与表头一致，保持统一 */" + \
                               "gridline-color: #ccc; /* 表格网格线颜色，与表头边框颜色相呼应 */" + \
                               "font-size: 12px; /* 数据字体比表头稍小一点，突出表头重要性 */" + \
                               "font-weight: normal; /* 正常字体粗细 */" + \
                               "}"

        self.update_port_button_style_sheet = "QPushButton {" + \
                                              "background-color: #1890FF; /* 按钮背景色为蓝色 */" + \
                                              "color: white; /* 按钮文本颜色为白色，与蓝色背景形成鲜明对比 */" + \
                                              "border: none; /* 无边框，使按钮看起来更简洁流畅 */" + \
                                              "padding: 6px 24px; /* 内边距，上下12px，左右12px，给文本足够空间 */" + \
                                              "text-align: center; /* 文本居中对齐，保证美观度 */" + \
                                              "text-decoration: none; /* 无文本装饰，保持简洁 */" + \
                                              "font-size: 16px; /* 字体大小为16px，清晰可读 */" + \
                                              "border-radius: 6px; /* 按钮四个角为半径6px的圆角，更显圆润精致 */" + \
                                              "}" + \
                                              "QPushButton:hover {" + \
                                              "background-color: #0070d8; /* 鼠标悬停时背景色变深一点 */" + \
                                              "}" + \
                                              "QPushButton:pressed {" + \
                                              "background-color: #0070d8; /* 按下时背景色再深一点 */" + \
                                              "}"

        self.start_test_button_style_sheet = "QPushButton {" + \
                                             "background-color: #1890FF; /* 按钮背景色为蓝色 */" + \
                                             "color: white; /* 按钮文本颜色为白色，与蓝色背景形成鲜明对比 */" + \
                                             "border: none; /* 无边框，使按钮看起来更简洁流畅 */" + \
                                             "padding: 12px 24px; /* 内边距，上下12px，左右24px，给文本足够空间 */" + \
                                             "text-align: center; /* 文本居中对齐，保证美观度 */" + \
                                             "text-decoration: none; /* 无文本装饰，保持简洁 */" + \
                                             "font-size: 20px; /* 字体大小为16px，清晰可读 */" + \
                                             "border-radius: 6px; /* 按钮四个角为半径6px的圆角，更显圆润精致 */" + \
                                             "}" + \
                                             "QPushButton:hover {" + \
                                             "background-color: #0070d8; /* 鼠标悬停时背景色变深一点 */" + \
                                             "}" + \
                                             "QPushButton:pressed {" + \
                                             "background-color: #0070d8; /* 按下时背景色再深一点 */" + \
                                             "}"
                                             
        self.pause_test_button_style_sheet = "QPushButton {" + \
                                            "background-color: #00FF00; /* 按钮背景色为绿色 */" + \
                                            "color: white; /* 按钮文本颜色为白色，与绿色背景形成鲜明对比 */" + \
                                            "border: none; /* 无边框，使按钮看起来更简洁流畅 */" + \
                                            "padding: 12px 24px; /* 内边距，上下12px，左右24px，给文本足够空间 */" + \
                                            "text-align: center; /* 文本居中对齐，保证美观度 */" + \
                                            "text-decoration: none; /* 无文本装饰，保持简洁 */" + \
                                            "font-size: 20px; /* 字体大小为16px，清晰可读 */" + \
                                            "border-radius: 6px; /* 按钮四个角为半径6px的圆角，更显圆润精致 */" + \
                                            "}" + \
                                            "QPushButton:hover {" + \
                                            "background-color: #00c200; /* 鼠标悬停时背景色变深一点 */" + \
                                            "}" + \
                                            "QPushButton:pressed {" + \
                                            "background-color: #00c200; /* 按下时背景色再深一点 */" + \
                                            "}"

        self.stop_test_button_style_sheet = "QPushButton {" + \
                                            "background-color: #FF0000; /* 按钮背景色为红色 */" + \
                                            "color: white; /* 按钮文本颜色为白色，与蓝色背景形成鲜明对比 */" + \
                                            "border: none; /* 无边框，使按钮看起来更简洁流畅 */" + \
                                            "padding: 12px 24px; /* 内边距，上下12px，左右24px，给文本足够空间 */" + \
                                            "text-align: center; /* 文本居中对齐，保证美观度 */" + \
                                            "text-decoration: none; /* 无文本装饰，保持简洁 */" + \
                                            "font-size: 20px; /* 字体大小为16px，清晰可读 */" + \
                                            "border-radius: 6px; /* 按钮四个角为半径6px的圆角，更显圆润精致 */" + \
                                            "}" + \
                                            "QPushButton:hover {" + \
                                            "background-color: #c20202; /* 鼠标悬停时背景色变深一点 */" + \
                                            "}" + \
                                            "QPushButton:pressed {" + \
                                            "background-color: #c20202; /* 按下时背景色再深一点 */" + \
                                            "}"

        # 定义样式表
        self.combo_box_style_sheet = "QComboBox {" + \
                                     "background-color: #f5f5f5; /* 淡灰色背景 */" + \
                                     "border: 1px solid #ccc; /* 浅灰色边框 */" + \
                                     "border-radius: 4px; /* 圆角 */" + \
                                     "padding: 5px; /* 内边距 */" + \
                                     "text-align: center; /* 使QComboBox显示框中的文本在水平方向居中 */" + \
                                     "font-size: 16px; /* 数据字体比表头稍小一点，突出表头重要性 */" + \
                                     "font-weight: bold; /* 正常字体粗细 */" + \
                                     "}" + \
                                     "QComboBox::down-arrow {" + \
                                     "image: url(arrow_down.png); /* 设置下拉箭头图片，可自行替换图片路径 */" + \
                                     "width: 12px; /* 箭头宽度 */" + \
                                     "height: 12px; /* 箭头高度 */" + \
                                     "}" + \
                                     "QComboBox QAbstractItemView::item {" + \
                                     "background-color: white; /* 下拉列表选项的白色背景 */" + \
                                     "color: #333; /* 选项文本颜色为深灰色 */" + \
                                     "padding: 5px; /* 选项内边距 */" + \
                                     "text-align: center; /* 使下拉列表中的选项文本在水平方向居中 */" + \
                                     "}"

    def init_widgets(self):
        self.submenu_load_scripts = self.window.findChild(QtWidgets.QAction, "submenu_load_scripts")
        self.submenu_load_scripts.triggered.connect(self.load_script)
        self.submenu_save_report = self.window.findChild(QtWidgets.QAction, "submenu_save_report")
        self.submenu_save_report.triggered.connect(self.save_report)
        self.submenu_exit = self.window.findChild(QtWidgets.QAction, "submenu_exit")
        self.submenu_version = self.window.findChild(QtWidgets.QAction, "submenu_version")
        self.submenu_version.triggered.connect(self.about_version)
        
        if self.log_ui_enable.lower() == 'y':
            self.submenu_log_ui = self.window.findChild(QtWidgets.QAction, "submenu_log_ui")
            self.submenu_log_ui.triggered.connect(self.show_log_ui)
            self.log_ui_window = uic.loadUi(uifile="ui/show_log.ui")
            self.log_ui_window.setWindowFlags(Qt.Window | Qt.CustomizeWindowHint | Qt.WindowTitleHint)
            self.textEdit_log = self.log_ui_window.findChild(QtWidgets.QTextEdit, "textEdit_log")
            self.stdout_redirector = self.StdoutRedirector(self.textEdit_log)
            stream_handler = logging.StreamHandler(stream=self.stdout_redirector)
            logger.addHandler(stream_handler)
            sys.stdout = self.stdout_redirector
            
        if self.current_ui_enable.lower() == 'y':
            self.submenu__current_ui = self.window.findChild(QtWidgets.QAction, "submenu__current_ui")
            self.motor_ui_window = uic.loadUi(uifile="ui/motor_current.ui")
            self.motor_ui_window.setWindowFlags(Qt.Window | Qt.CustomizeWindowHint | Qt.WindowTitleHint)
            self.submenu__current_ui.triggered.connect(self.toggle_current_ui)
            self.init_current_ui_widgets()
        
        self.lbl_promt = self.window.findChild(QtWidgets.QLabel, "lbl_promt")
        self.lbl_promt.setVisible(False)
        self.lbl_promt.setStyleSheet("color: blue;font-size: 18px")
        self.cbx_aging_time = self.window.findChild(QtWidgets.QComboBox, "cbx_aging_time")
        self.cbx_aging_time.addItems(self.aging_duration_options)
        self.cbx_aging_time.setStyleSheet(self.combo_box_style_sheet)
        self.cbx_aging_time.currentIndexChanged.connect(self.on_aging_cbx_changed)
        self.lbl_remain_time = self.window.findChild(QtWidgets.QLabel, "lbl_remain_time")
        self.lbl_remain_time.setStyleSheet("font-size: 18px;font-weight: bold")

        self.chb_com_all = self.window.findChild(QtWidgets.QCheckBox, "chb_com_all")
        self.chb_com_all.clicked.connect(lambda checked, cb=self.chb_com_all: self.on_port_cbx_clicked(checked, cb))
        self.port_Layout = self.window.findChild(QtWidgets.QHBoxLayout, "port_Layout")
        
        self.btn_fresh_ports = self.window.findChild(QtWidgets.QPushButton, "btn_fresh_ports")
        self.btn_fresh_ports.setStyleSheet(self.update_port_button_style_sheet)
        self.btn_fresh_ports.clicked.connect(self.refresh_ports)

        self.tv_test_detail = self.window.findChild(QtWidgets.QTableView, "tableView")
        self.set_model()

        self.btn_start_test = self.window.findChild(QtWidgets.QPushButton, "btn_start_test")
        self.btn_start_test.setStyleSheet(self.start_test_button_style_sheet)
        self.btn_start_test.clicked.connect(self.start_test)
        
        self.btn_pause_test = self.window.findChild(QtWidgets.QPushButton, "btn_pause_test")
        self.btn_pause_test.setStyleSheet(self.pause_test_button_style_sheet)
        self.btn_pause_test.clicked.connect(self.pause_test)

        self.btn_stop_test = self.window.findChild(QtWidgets.QPushButton, "btn_stop_test")
        self.btn_stop_test.setStyleSheet(self.stop_test_button_style_sheet)
        self.btn_stop_test.clicked.connect(self.stop_test)
        
        self.update_port_options(startup=True)
        
    def init_current_ui_widgets(self):
        label_names = [f"label_com{i}" for i in range(1, 17)] + \
                    [f"text_current_{i}" for i in range(11, 17)] + \
                    [f"text_current_{i}" for i in range(21, 27)] + \
                    [f"text_current_{i}" for i in range(31, 37)] + \
                    [f"text_current_{i}" for i in range(41, 47)] + \
                    [f"text_current_{i}" for i in range(51, 57)] + \
                    [f"text_current_{i}" for i in range(61, 67)] + \
                    [f"text_current_{i}" for i in range(71, 77)] + \
                    [f"text_current_{i}" for i in range(81, 87)] + \
                    [f"text_current_{i}" for i in range(91, 97)] + \
                    [f"text_current_{i}" for i in range(101, 107)] + \
                    [f"text_current_{i}" for i in range(111, 117)] + \
                    [f"text_current_{i}" for i in range(121, 127)] + \
                    [f"text_current_{i}" for i in range(131, 137)] + \
                    [f"text_current_{i}" for i in range(141, 147)] + \
                    [f"text_current_{i}" for i in range(151, 157)] + \
                    [f"text_current_{i}" for i in range(161, 167)]

        for name in label_names:
            widget = self.motor_ui_window.findChild(QtWidgets.QWidget, name)
            if isinstance(widget, QtWidgets.QLabel):
                self.label_com_list.append(widget)
            elif isinstance(widget, QtWidgets.QTextEdit):
                self.editText_current_list.append(widget)
                
    def update_current_ui_portnames(self,ports=[]):
        if ports[0] == self.no_used_port:
            for label in self.label_com_list:
                label.setEnabled(False)
            for editText in self.editText_current_list:
                editText.setEnabled(False)
        else:
            num_labels_to_show = min(len(ports), len(self.label_com_list))
            num_edit_texts_to_show = num_labels_to_show * 6

            for i in range(num_labels_to_show):
                label_text = f"{ports[i]}"
                self.label_com_list[i].setText(label_text)
                self.label_com_list[i].setEnabled(True)
                
            for i in range(num_edit_texts_to_show):
                self.editText_current_list[i].setEnabled(True)

            for i in range(num_labels_to_show, len(self.label_com_list)):
                self.label_com_list[i].setVisible(False)

            for i in range(num_edit_texts_to_show, len(self.editText_current_list)):
                self.editText_current_list[i].setVisible(False)
                
    def update_current_ui_motorcurrents(self,currents={}):
        # 遍历整理好的数据，将对应数据赋值给editText_current_list中的editText
        for port, data_list in currents.items():
            if port in self.port_names:
                port_index = self.port_names.index(port)
                start_edit_text_index = port_index * 6

                for data in data_list:
                    thumb_root_text = f"{data.get('thumb_root', '')}"
                    self.editText_current_list[start_edit_text_index].setText(thumb_root_text)

                    thumb_text = f"{data.get('thumb', '')}"
                    self.editText_current_list[start_edit_text_index + 1].setText(thumb_text)

                    index_text = f"{data.get('index', '')}"
                    self.editText_current_list[start_edit_text_index + 2].setText(index_text)

                    middle_text = f"{data.get('middle', '')}"
                    self.editText_current_list[start_edit_text_index + 3].setText(middle_text)

                    third_text = f"{data.get('third', '')}"
                    self.editText_current_list[start_edit_text_index + 4].setText(third_text)

                    little_text = f"{data.get('little', '')}"
                    self.editText_current_list[start_edit_text_index + 5].setText(little_text)
            
            
    def show_log_ui(self):
        """显示日志界面的方法"""
        if self.submenu_log_ui.isChecked():
            self.log_ui_window.show()
        else:
            self.log_ui_window.hide()

    def toggle_current_ui(self):
        """切换当前界面显示状态的方法"""
        if self.submenu__current_ui.isChecked():
            self.motor_ui_window.show()
        else:
            self.motor_ui_window.hide()

    def set_model(self):
        self.model = QStandardItemModel()
        self.model.setHorizontalHeaderLabels(self.HEADS)
        self.tv_test_detail.setModel(self.model)
        self.tv_test_detail.setStyleSheet(self.str_header_format + self.str_data_format)
        # self.tv_test_detail.horizontalHeader().setSectionsClickable(True)
        delegate = CustomDelegate(self.tv_test_detail)
        for column in range(self.model.columnCount()):
            self.tv_test_detail.setItemDelegateForColumn(column, delegate)
    
    def count_down(self, hour=0.5):
        if not self.timer_running:
            self.start_time = datetime.datetime.now()
            self.base_time = self.start_time
            self.end_time = self.start_time + datetime.timedelta(hours=float(hour) + float(self.offset_duration))
            self.timer = QtCore.QTimer(self)
            self.timer.timeout.connect(self.update_remaining_time)
            self.timer.start(1000)
            self.timer_running = True

    def update_remaining_time(self):
        current_time = datetime.datetime.now()
        time_difference = self.end_time - current_time
        if time_difference.total_seconds() > 0:
            total_seconds = time_difference.total_seconds()
            hours = int(total_seconds // 3600)
            remaining_seconds = total_seconds % 3600
            minutes = int(remaining_seconds // 60)
            seconds = int(remaining_seconds % 60)
            self.lbl_remain_time.setText(f"{hours:02d}:{minutes:02d}:{seconds:02d}")
        else:
            self.timer.stop()
            self.timer_running = False
            self.lbl_remain_time.setText("已结束")

    def on_aging_cbx_changed(self, index):
        self.selected_aging_duration = self.cbx_aging_time.itemText(index)
        self.offset_duration = self.get_offset_duration()
        logger.info(f'aging 选项为：{self.selected_aging_duration}')

    def on_port_cbx_clicked(self, checked, checkbox):
        """
        当QCheckBox被点击时调用的函数
        :param checked: 表示QCheckBox是否被选中的状态
        :param checkbox: 被点击的QCheckBox对象
        """
        port = checkbox.text()
        if port == '全选':
            self.select_port_names = self.port_names
            for cbx in self.check_box_list:
                cbx.setChecked(checked)
                port2 = cbx.text()
                self.update_device_list(port=port2, isChecked=checked)
        else:
            self.select_port_names.append(port)
            self.update_device_list(port=port, isChecked=checked)

    def remove_all_widgets_from_layout(self, layout):
        while layout.count():
            item = layout.takeAt(0)
            if item.widget():
                item.widget().setParent(None)
            elif item.layout():
                self.remove_all_widgets_from_layout(item.layout())

    def update_port_options(self, startup=False):
        """
        在给定的水平布局中根据QCheckBox的数量动态添加垂直布局，并将QCheckBox合理分配到各个垂直布局中，
        确保生成的QCheckBox行和列分别对齐且每一列都是8个，同时设置列间距为3像素，行间距为8像素。
        :param layout: 传入的水平布局对象
        """
        vertical_layouts = []  # 用于存储创建的垂直布局对象
        q_check_box_count = 0  # 用于计数已添加的QCheckBox数量

        if startup:
            # 启动时清理布局及相关列表，添加提示标签
            self.remove_all_widgets_from_layout(self.port_Layout)
            self.check_box_list.clear()
            self.set_checked_box_status(False)
            startup_label = QLabel('请先刷新端口，获取设备信息')
            self.port_Layout.addWidget(startup_label)
            logger.info('update_port_options,fist startup')
            return

        logger.info('update_port_options,start collect port infos')
        # 显示提示信息，清理布局及相关数据结构
        self.lbl_promt.setVisible(True)
        self.remove_all_widgets_from_layout(self.port_Layout)
        self.devices_info_list.clear()
        self.model.clear()
        self.model.setHorizontalHeaderLabels(self.HEADS)

        # 启动线程获取端口信息，添加异常处理
        try:
            threading.Thread(target=self.get_port_infos).start()
        except Exception as e:
            logger.error(f"Failed to start the thread for getting port infos: {e}")

        # 等待端口信息收集完成，期间适时处理事件以保持界面响应，优化等待机制（这里只是示例思路，可根据实际情况细化）
        wait_start_time = time.time()
        while True:
            if not self.update_port_enable:
                elapsed_time = time.time() - wait_start_time
                if elapsed_time > self.time_out:  # 设置一个超时时间
                    logger.error("Timeout while waiting for port infos collection.")
                    self.update_port_enable = True
                    break
                time.sleep(1)  # 缩短等待时间，动态调整可根据实际情况
                wait_str = f'设备信息读取中，请耐心等待......{self.time_out-int(elapsed_time)}秒'
                self.lbl_promt.setText(wait_str)
                self.app.processEvents()
            else:
                logger.info(f'update_port_options:collect port infos completely')
                self.remove_all_widgets_from_layout(self.port_Layout)
                break

        self.lbl_promt.setVisible(False)

        # 根据是否有可用端口进行不同处理
        if self.port_names[0] == self.no_used_port:
            no_port_label = QLabel(self.port_names[0])
            self.port_Layout.addWidget(no_port_label)
            return
        self.set_checked_box_status(True)
        # 遍历端口名称，创建并添加QCheckBox到合适的垂直布局
        for port in self.port_names:
            check_box = QCheckBox(port)
            self.check_box_list.append(check_box)

            q_check_box_count += 1

            column_index = (q_check_box_count - 1) // 8
            if column_index < 4:
                if len(vertical_layouts) <= column_index:
                    vertical_layout = QVBoxLayout()
                    self.port_Layout.addLayout(vertical_layout)
                    vertical_layouts.append(vertical_layout)

                vertical_layouts[column_index].addWidget(check_box)
                vertical_layouts[column_index].setAlignment(Qt.AlignTop)
                # 为每个QCheckBox添加响应函数，优化为使用命名函数包裹
                def on_port_cbx_clicked_wrapper(checked, cb):
                    return self.on_port_cbx_clicked(checked, cb)
                check_box.clicked.connect(lambda checked, cb=check_box: on_port_cbx_clicked_wrapper(checked, cb))

        # 设置列间距为18像素
        self.port_Layout.setContentsMargins(0, 0, 18, 0)

        # 设置行间距为25像素
        for vertical_layout in vertical_layouts:
            vertical_layout.setSpacing(25)

        self.port_Layout.setSpacing(10)
        self.update_port_enable = True
        self.last_refresh_time = self.current_time
        
    def set_checked_box_status(self,enable=False):
        for checkbox in self.check_box_list:
            checkbox.setEnabled(enable)
        self.chb_com_all.setEnabled(enable)
        # self.btn_fresh_ports.setEnabled(enable)
        self.cbx_aging_time.setEnabled(enable)

    def start_test(self):
        logger.info('start_test')
        if self.port_names[0] == self.no_used_port:
            logger.error(self.no_used_port)
            QMessageBox.information(self.window, '提示', f"无可用端口,请先刷新端口后再做尝试")
            return
        
        if self.model.rowCount() < 1:
            logger.error('请添加测试设备')
            QMessageBox.information(self.window, '提示', f"请添加测试设备")
            return
        
        self.execute_script ()
        self.write_to_json_file(stop_test=False,pause_test=False)

    def execute_script (self):
        if self.script_name is not None:
            if not self.running:
                result = QMessageBox.question(self.window, '确认', '测试即将开始，请耐心等待。是否继续执行？',
                                              QMessageBox.Yes | QMessageBox.No)
                if result == QMessageBox.Yes:
                 
                    # 将模块所在的父目录（即scripts所在目录）添加到sys.path中，确保能找到模块
                    module_dir = os.path.dirname(self.script_name)
                    sys.path.append(module_dir)
                    try:
                        # 尝试导入脚本模块，先获取去掉扩展名后的模块名部分
                        module_name = os.path.splitext(os.path.basename(self.script_name))[0]
                        module = importlib.import_module(module_name)
                        self.run_script_thread = threading.Thread(target=self.update_test_result, args=(module,))
                        self.run_script_thread.start()
                        self.running = True
                        self.set_checked_box_status(False)
                        self.update_device_Info_worker = self.UpdateDeviceInfoWorker()
                        self.update_device_Info_worker.selected_aging_duration = self.selected_aging_duration
                        self.update_device_Info_worker.off_duration = self.offset_duration
                        self.update_device_Info_worker.test_result = self.get_test_result()
                        self.update_device_Info_worker.update_progress_signal.connect(self.update_device_info_progress)
                        self.update_device_Info_worker.update_result_signal.connect(self.update_device_info_result)
                        self.update_device_Info_thread = threading.Thread(target=self.update_device_Info_worker.run_test)
                        self.update_device_Info_thread.start()
                        self.count_down(hour=self.selected_aging_duration)
                    except ImportError as e:
                        result = QMessageBox.critical(self.window, '错误',
                                                      f"导入模块失败：{self.script_name}，错误信息：{e}")
                        return
                else:
                    # 用户选择了“否”，不执行
                    self.running = False
                    return
            else:
                result = QMessageBox.information(self.window, '提示', f"有任务在运行...不要重复加载任务")
        else:
            result = QMessageBox.information(self.window, '提示', f"请先加载脚本")
            self.running = False

    def update_test_result(self, module):
        def run_script():
            try:
                self.report_title,self.overall_result,self.need_show_current = module.main(ports=self.select_port_names,node_ids=self.node_ids,
                                                     aging_duration=float(self.selected_aging_duration))
                logger.info(f'本次测试已结束，详细测试数据为：\n')
                
                self.update_device_Info_worker.update_test_result()
                self.print_overall_result(self.overall_result)
                if self.need_show_current:
                    test_data = self.get_currents_from_test_result(self.overall_result)
                    self.update_current_ui_worker = self.UpdateCurrentUIWorker(portName=self.port_names,test_data = test_data)
                    if self.current_ui_enable.lower() == 'y':
                        self.update_current_ui_worker.update_com_name_signal.connect(self.update_current_ui_portnames)
                        self.update_current_ui_worker.update_current_signal.connect(self.update_current_ui_motorcurrents)
                        self.update_current_ui_worker.run_test()
                self.running = False
                self.set_checked_box_status(True)
                self.update_device_Info_worker.test_result = self.get_test_result()
            except Exception as e:
                logger.error(f'Error in script execution: {e}')

        # 异步更新界面
        thread = threading.Thread(target=run_script)
        thread.daemon = True  # 主界面退出，子任务也能退出
        thread.start()
        
    def get_currents_from_test_result(self, overall_result):
        port_data_dict = {}

        # 整理数据
        for item in overall_result:
            if item['port'] not in port_data_dict:
                port_data_dict[item['port']] = []
            for gesture in item['gestures']:
                port_data_dict[item['port']].append(gesture['content'])
                
        return port_data_dict

    def print_overall_result(self, overall_result):
        port_data_dict = {}

        # 整理数据
        for item in overall_result:
            if item['port'] not in port_data_dict:
                port_data_dict[item['port']] = []
            for gesture in item['gestures']:
                port_data_dict[item['port']].append((gesture['timestamp'],gesture['description'],gesture['expected'],gesture['content'], gesture['result'], gesture['comment']))

        # 打印数据
        for port, data_list in port_data_dict.items():
            logger.info(f"Port: {port}")
            for timestamp, description, expected, content, result, comment in data_list:
                logger.info(f" timestamp:{timestamp} ,description:{description},expected:{expected},content: {content}, Result: {result},comment:{comment}")

    def update_device_info_progress(self, percentage):
        for port in self.select_port_names:
            self.update_device_info(port=port, key=self.STR_TEST_PROGRESS, new_value=f"{percentage:.2f}%")

    def update_device_info_result(self, result):
        if result is not None and len(result)> 0:
            for port, result in result.items():
                self.update_device_info(port=port, key=self.STR_TEST_RESULT, new_value=result)

    def update_device_info(self, port, key, new_value):
        for i, device_info in enumerate(self.devices_info_list):
            if device_info[self.STR_PORT] == port:
                device_info[key] = new_value
                # 更新视图模型
                for j, header in enumerate(self.HEADS):
                    if header == key:
                        item = self.model.item(i, j)
                        item.setText(str(new_value))
        self.tv_test_detail.update()

    def stop_test(self):
        logger.info('stop_test')
        if self.running:
            if self.timer_running:
                self.timer.stop()
                self.timer_running = False

            # 假设存在与测试线程相关的操作，以下是完善后的处理方式
            if hasattr(self, 'update_device_Info_thread') and self.update_device_Info_thread.is_alive():
                self.update_device_Info_worker.stop_flag=True
                self.update_device_Info_worker.test_finished_signal.connect(self.on_test_finished)
                self.run_script_thread.join()
            else:
                self.on_test_finished()
            
            self.running = False
            self.write_to_json_file(stop_test=True, pause_test=True)
        
    def pause_test(self):
        pause = False
        if self.running:
            _,pause = self.read_from_json_file()
            if not pause:
                pause = True
                self.btn_pause_test.setText('恢复测试')
                self.update_device_Info_worker.pause_flag = True
                # self.update_device_Info_worker.test_result = '暂停测试'
                logger.info('pause test')
            else:
                pause = False
                self.btn_pause_test.setText('暂停测试')
                self.update_device_Info_worker.pause_flag = False
                # self.update_device_Info_worker.test_result = '进行中'
                logger.info('go on  test')
                
            self.write_to_json_file(stop_test=False, pause_test=pause)
            
            
    def on_test_finished(self):
        # 恢复相关界面控件的可操作性
        self.set_checked_box_status(True)
        self.btn_fresh_ports.setEnabled(True)
        self.cbx_aging_time.setEnabled(True)


    def refresh_ports(self):
        self.current_time = time.time()
        if (self.current_time - self.last_refresh_time >= 5) and not self.running and self.update_port_enable:
            logger.info('start refresh ports')
            self.update_port_enable = False
            # threading.Thread(target=self.update_port_options).start()
            self.update_port_options()
            self.update_current_ui_portnames(ports=self.port_names)
        else:
            logger.info(' do not refresh')

    def load_script(self):
        logger.info('start load script')
        # 弹出文件选择对话框，让用户选择要执行的脚本
        file_path, _ = QFileDialog.getOpenFileName(self.window, '选择要执行的脚本', 'scripts', 'Python files (*.py)')
        if file_path:
            # 获取文件名（包含扩展名）
            file_name = os.path.basename(file_path)
            # 获取scripts目录的绝对路径（确保后续拼接路径的准确性）
            scripts_dir = os.path.abspath('scripts')
            # 使用with语句打开脚本文件，确保在读取完成后自动关闭文件,释放资源
            with open(file_path, 'r') as f:
                # 直接使用os.path.join将scripts（绝对路径形式）和文件名拼接起来
                self.script_name = os.path.join(scripts_dir, file_name)
                logger.info(f'加载脚本{self.script_name}，请点击开始测试按钮，执行脚本\n')
                
    def get_test_result(self):
        port_result_dict = {}
        if  not self.running and len(self.overall_result)>0:
            try:
                for item in self.overall_result:
                    port = item['port']
                    if port not in port_result_dict:
                        port_result_dict[port] = '通过' 
                    for gesture in item['gestures']:
                        result = gesture['result']
                        if result == '不通过':
                            port_result_dict[port] = '不通过'  # 只要有一个不通过，端口结果设为不通过
                            break  # 一旦发现不通过，无需再检查该端口下其他手势结果，直接跳出内层循环
                return port_result_dict
            except KeyError as e:
                logger.error(f"数据结构中缺少关键键值对，异常信息: {e}")
                return {}  # 如果出现异常，返回空字典
        else:
            return {}
        # else:
        #     for port in self.port_names:
        #         port_result_dict[port] = '进行中'
        #     return port_result_dict
                
    def extract_test_data (self):
        port_data_dict = {}

        # 整理数据
        for item in self.overall_result:
            if item['port'] not in port_data_dict:
                port_data_dict[item['port']] = []
            for gesture in item['gestures']:
                port_data_dict[item['port']].append((gesture['timestamp'],gesture['description'],gesture['expected'],gesture['content'], gesture['result'], gesture['comment']))
        return port_data_dict
    
    def save_report(self):
        # 表头
        headers = ["用例编号", "用例描述", "期望值", "实际值", "是否通过", "备注"]

        # 表格宽高度
        column_widths = [10, 30, 12, 12, 10, 11]
        default_row_height = 35  # 设置默认行高为18

        # 设置表格边框样式
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        title_fill = PatternFill(start_color="B2DFEE", end_color="B2DFEE", fill_type="solid")
        title_fill_font_color = Font(color="333333")

        # 单元格对齐方式（居中对齐且自动换行）
        alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # 表头填充颜色（淡蓝色）
        header_fill = PatternFill(start_color="B2DFEE", end_color="B2DFEE", fill_type="solid")
        header_font_color = Font(color="333333")

        # 奇数行填充颜色（淡米色）
        odd_row_fill = PatternFill(start_color="FDF5E6", end_color="FDF5E6", fill_type="solid")
        # 偶数行填充颜色（淡灰色）
        even_row_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

        # 创建工作簿
        wb = Workbook()
        ws = wb.active

        # 设置标题占两列并居中
        report_cell = ws.merge_cells('A1:F1')
        ws['A1'] = self.report_title
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = alignment
        ws['A1'].fill = title_fill
        ws['A1'].font = title_fill_font_color

        ws['A2'] = '产品名称'
        ws['C2'] = '产品型号'
        ws.merge_cells('D2:F2')

        ws['A3'] = '测试地点'
        ws['C3'] = '测试时间'
        ws.merge_cells('D3:F3')

        ws['A4'] = '测试设备'
        ws['C4'] = '设备编号'
        ws.merge_cells('D4:F4')

        ws['A5'] = '温度'
        ws['C5'] = '湿度'
        ws.merge_cells('D5:F5')

        ws.append(headers)

        # 设置表头字体为粗体，修正索引为2（对应第2行，表头所在行）
        for cell in ws[6]:
            cell.fill = header_fill
            cell.font = header_font_color

        # 设置默认行高
        for row in ws.iter_rows(min_row=ws.min_row, max_row=ws.max_row-1):
            ws.row_dimensions[row[0].row].height = default_row_height

        # 测试数据填充
        self.test_data = self.extract_test_data ()
        row_index = 7  # 从第7行开始填充数据行
        
        for port, data_list in self.test_data.items():
            for timestamp, description, expected, content, result, comment in data_list:
                row_data = [row_index-6, description, str(expected), str(content), result,
                            f"[{port}]{timestamp} {comment}"]
                ws.append(row_data)
                if row_index % 2 == 1:
                    fill_color = odd_row_fill
                else:
                    fill_color = even_row_fill
                for cell in ws[row_index]:
                    cell.fill = fill_color
                row_index += 1

        # 设置自动换行后自适应行高（直接设置height为None来实现自适应）
        for row in ws.iter_rows(min_row=7, max_row=ws.max_row):
            row_dimension = RowDimension(ws, row[0].row)
            row_dimension.height = None  # 设置为None让其根据内容自适应高度
            ws.row_dimensions[row[0].row] = row_dimension

        # ws.merge_cells('E2:E3')
        ws['A' + str(row_index)] = '测试结论'
        ws.merge_cells('B' + str(row_index) + ':' + 'F' + str(row_index))
        ws['B' + str(row_index)] = '[ ]合格               [ ]不合格'

        ws['A' + str(row_index + 1)] = '测试人'
        ws['C' + str(row_index + 1)] = '审核人'
        ws.merge_cells('D' + str(row_index + 1) + ':' + 'F' + str(row_index + 1))

        ws['A' + str(row_index + 2)] = '备注'
        ws.merge_cells('B' + str(row_index + 2) + ':' + 'F' + str(row_index + 2))
        
        for row in ws.iter_rows(min_row=row_index, max_row=ws.max_row):
             ws.row_dimensions[row[0].row].height = default_row_height

        # 设置列宽并统一设置对齐方式（避免重复设置对齐方式）
        for i in range(len(column_widths)):
            col_letter = get_column_letter(i + 1)
            ws.column_dimensions[col_letter].width = column_widths[i]
        for row in ws.rows:
            for cell in row:
                cell.alignment = alignment

        # 设置表格边框样式
        for row in ws.iter_rows(min_row=1, max_row=len(ws['A']), min_col=1, max_col=6):
            for cell in row:
                cell.border = thin_border

        # 根据时间动态生成文件名，避免覆盖（示例格式，可根据需求调整）
        current_dir = os.getcwd()
        # 拼接出log文件夹的路径
        report_folder_path = os.path.join(current_dir, "report")
        # 判断log文件夹是否存在，如果不存在则创建它，添加异常处理
        try:
            if not os.path.exists(report_folder_path):
                os.mkdir(report_folder_path)
        except OSError as e:
            logger.error(f"创建文件夹 {report_folder_path} 时出错: {e}")
            return  # 或者可以采取其他处理方式，比如提示用户手动创建文件夹后重新运行等
        if self.script_name is not None:
            name = os.path.splitext(os.path.basename(self.script_name))[0]
            file_name = f"{name}_report_{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
            file_path = os.path.join(report_folder_path, file_name)

            # 保存工作簿
            wb.save(file_path)
            QMessageBox.information(self.window, '保存成功', f'测试报告已保存为：{file_path}')
        else:
            QMessageBox.information(self.window, '保存失败', f'尚未做测试，请先进行测试')
    
                
    def about_version(self):
        """
        显示关于版本信息的弹出窗口。

        这个函数创建一个弹出窗口，显示软件版本、发布时间和版权信息。
        """
        text = f"软件版本: {self.client_version}\n发布时间: {self.release_date_str}\n版权所有© 2015·2024 上海傲意信息科技有限公司"
        QMessageBox.information(self.window,'软件版本', text)
        
    def get_port_infos(self):
        """
        通过并行任务的方式获取端口相关设备信息，提高获取效率。
        """
        portInfos = serial.tools.list_ports.comports()
        ports = [portInfo.device for portInfo in portInfos if portInfo]
        portNames = []
        nodeIds = []
        MAX_NODE_ID = 256
        ROH_FW_VERSION = 1001  # 固件版本寄存器地址

        def process_port(port):
            """
            处理单个端口信息获取的函数，每个端口会在独立的线程中执行此函数。
            """
            node_id = '无法获取'
            sw_version = '无法获取'
            connect_status = '未连接'
            client = ModbusClient(port=port)
            try:
                client.connect()
                for id in range(2,MAX_NODE_ID):
                    logger.info(f'check port device id:{id}')
                    response = client.serialclient.read_holding_registers(ROH_FW_VERSION, 2, id)
                    if not response.isError():
                        portNames.append(port)
                        nodeIds.append(id)
                        sw_version = self.convert_version_format(response)
                        node_id = id
                        connect_status = '已连接'
                        devices_info = {
                            self.STR_PORT: port,
                            self.STR_DEVICE_NAME: 'Rohand',
                            self.STR_SOFTWARE_VERSION: sw_version,
                            self.STR_DEVICE_ID: node_id,
                            self.STR_CONNECT_STATUS: connect_status,
                            self.STR_TEST_PROGRESS: '0%',
                            self.STR_TEST_RESULT: '--'
                        }
                        return devices_info
                client.disconnect()
                return None
            except Exception as e:
                logger.error(f"Error during setup for port {port}: {e}\n")
                return None
            finally:
                if client:
                    client.dis_connect()

        # 使用线程池并发执行每个端口的任务
        with concurrent.futures.ThreadPoolExecutor() as executor:
            results = list(executor.map(process_port, ports))

        # 处理收集到的结果，提取有效设备信息并更新类属性
        valid_devices_info = [info for info in results if info is not None]
        if valid_devices_info:
            self.port_names = [info[self.STR_PORT] for info in valid_devices_info]
            self.node_ids = [info[self.STR_DEVICE_ID] for info in valid_devices_info]
            self.devices_info_list = valid_devices_info
        else:
            self.port_names = [self.no_used_port]
            self.node_ids = [2]

        self.update_port_enable = True
            

    def convert_version_format(self, response):
        """
        从给定的响应中提取版本号并转换为特定格式。

        参数：
        response：包含从寄存器读取的值的对象。

        返回：
        以“V主版本号.次版本号.补丁版本号”格式的字符串版本号
        """
        if hasattr(response, 'registers'):
            if len(response.registers) > 0:
                value1 = response.registers[0]
                value2 = response.registers[1]
                major_version = (value1 >> 8) & 0xFF
                minor_version = value1 & 0xFF
                patch_version = value2 & 0xFF
                return f"V{major_version}.{minor_version}.{patch_version}"
        else:
            return "无法获取"

    def update_device_info(self, port, key, new_value):
        for i, device_info in enumerate(self.devices_info_list):
            if device_info[self.STR_PORT] == port:
                device_info[key] = new_value
                # 更新视图模型
                for j, header in enumerate(self.HEADS):
                    if header == key:
                        if i < self.model.rowCount() and j < self.model.columnCount():
                            item = self.model.item(i, j)
                            item.setText(str(new_value))
                        else:
                            logger.error(f"行索引 {i} 或列索引 {j} 超出了模型的有效范围。")
                        # # 发出dataChanged信号，通知视图数据已改变
                        # topLeftIndex = self.model.index(i, j)
                        # bottomRightIndex = self.model.index(i, j)
                        # self.model.dataChanged(topLeftIndex, bottomRightIndex)
        self.tv_test_detail.update()

    def update_device_list(self, port, isChecked):
        if len(self.devices_info_list)==0:
            return
        
        if isChecked:  # 新增数据
            device_info = self.get_device_Info(port)
            new_item_data = []
            for header in self.HEADS:
                value = device_info.get(header)
                if value is not None:
                    item = QStandardItem(str(value))
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    new_item_data.append(item)

            # 检查model中是否已存在相同port的数据
            port_key_index = self.HEADS.index(self.STR_PORT)
            is_duplicate = False
            for row in range(self.model.rowCount()):
                existing_item = self.model.item(row, port_key_index)
                if existing_item and existing_item.text() == device_info.get(self.STR_PORT):
                    is_duplicate = True
                    break

            if not is_duplicate:
                self.model.appendRow(new_item_data)

        else:  # 删除数据
            row_to_delete = None
            for row in range(self.model.rowCount()):
                device_info = {}
                for col in range(self.model.columnCount()):
                    item = self.model.item(row, col)
                    header = self.HEADS[col]
                    device_info[header] = item.text()

                if device_info.get(self.STR_PORT) == port:
                    row_to_delete = row
                    break

            if row_to_delete is not None:
                # 从model中删除指定行
                self.model.removeRow(row_to_delete)

    def get_device_Info(self, port):
        devices_info = {}
        for device in self.devices_info_list:
            if device.get(self.STR_PORT) == port:
                devices_info = device
                break
       
        return devices_info
    
    class UpdateCurrentUIWorker(QObject):
        update_com_name_signal = pyqtSignal(list)
        update_current_signal = pyqtSignal(dict)
        
        def __init__(self, portName, test_data, parent=None):
            super().__init__(parent)
            self.portName = portName
            self.test_data = test_data
            
        def run_test(self):
            self.update_com_name_signal.emit(self.portName)
            self.update_current_signal.emit(self.test_data)
        
    class UpdateDeviceInfoWorker(QObject):
        update_progress_signal = pyqtSignal(float)
        update_result_signal = pyqtSignal(dict)
        test_finished_signal = pyqtSignal()  # 添加这个信号

        def __init__(self, parent=None):
            super().__init__(parent)
            self.stop_flag = False
            self.pause_flag = False
            self.selected_aging_duration = None
            self.off_duration = 0
            self.test_result = {}
            
        def run_test(self):
            start_time = datetime.datetime.now()
            end_time = start_time + datetime.timedelta(hours=float(self.selected_aging_duration)+float(self.off_duration))
            self.update_result_signal.emit(self.test_result)
            while datetime.datetime.now() < end_time:
                if self.stop_flag:
                    break
                if self.pause_flag:
                    self.update_result_signal.emit(self.test_result)
                    continue
                self.update_result_signal.emit(self.test_result)
                elapsed_time = datetime.datetime.now() - start_time
                percentage = (elapsed_time.total_seconds() / (((float(self.selected_aging_duration)+float(self.off_duration)) * 3600))) * 100
                self.update_progress_signal.emit(percentage)
                # 这里可以添加适当的时间间隔控制，避免过于频繁地更新信号发送
                # import time
                time.sleep(1)
            # 最后将测试进度更新为100%
            self.update_result_signal.emit(self.test_result)
            self.update_progress_signal.emit(100)
            self.test_finished_signal.emit()  
            
        def update_test_result(self):
            self.update_result_signal.emit(self.test_result)
                
class CustomDelegate(QtWidgets.QStyledItemDelegate):
    def paint(self, painter, option, index):
        option = QtWidgets.QStyleOptionViewItem(option)
        self.initStyleOption(option, index)
        item = index.model().itemFromIndex(index)
        column_name = index.model().headerData(index.column(), Qt.Horizontal, Qt.DisplayRole)
        value = item.text()
        # 设置文本居中对齐
        option.displayAlignment = Qt.AlignmentFlag.AlignCenter
         # 创建字体对象并设置字体及大小属性，这里示例设置为字体 "Arial"，大小为12
        font = QFont("Arial", 10)
        font.setBold(True)  # 设置字体加粗
        option.font = font

        if column_name == ClientTest.STR_CONNECT_STATUS:
            if value == "已连接":
                option.palette.setColor(QtGui.QPalette.Text, QColor("green"))
            elif value == "未连接":
                option.palette.setColor(QtGui.QPalette.Text, QColor("gray"))

        elif column_name == ClientTest.STR_TEST_PROGRESS:
            try:
                percentage = float(value.rstrip('%'))  # 去除百分号并转换为浮点数
            except (ValueError, AttributeError):
                percentage = 0.0
            rect = option.rect
            bar_rect = rect.adjusted(2, 2, -2, -2)  # 调整进度条绘制区域，留出边框空间
            painter.save()
            painter.setRenderHint(QPainter.Antialiasing)
            # 绘制进度条背景
            painter.setPen(Qt.NoPen)
            painter.setBrush(QColor("lightgray"))
            painter.drawRect(bar_rect)
            # 绘制进度部分
            progress_rect = bar_rect.adjusted(0, 0, round(-(bar_rect.width() * (1 - percentage / 100))), 0)
            if percentage > 0 and percentage < 100:
                # painter.setBrush(QColor("blue"))
                painter.setBrush(QColor("#1890FF"))
            else:
                # painter.setBrush(QColor("green"))
                painter.setBrush(QColor("#73CD73"))
            painter.drawRect(progress_rect)
            painter.restore()
        elif column_name == ClientTest.STR_TEST_RESULT:
            if value == "通过":
                option.palette.setColor(QtGui.QPalette.Text, QColor("green"))
            elif value == "不通过":
                option.palette.setColor(QtGui.QPalette.Text, QColor("red"))
            else:
                option.palette.setColor(QtGui.QPalette.Text, QColor("black"))

        super().paint(painter, option, index)
        
class ModbusClient:
    baudrate=115200
    framer=FramerType.RTU
    port = None
    serialclient = None
    
    def __init__(self,port):
        self.port = port
        
    def connect(self):
        try:
            self.serialclient = ModbusSerialClient(self.port, self.framer, self.baudrate)
            if not self.serialclient.connect():
                raise ConnectionException(f"[port = {self.port}]Could not connect to Modbus device.")
            logger.info(f"[port = {self.port}]Successfully connected to Modbus device.")
        except ConnectionException as e:
            logger.error(f"[port = {self.port}]Error during connection: {e}")
            raise
        
    def dis_connect(self):
       if self.serialclient:
            try:
                self.serialclient.close()
                self.serialclient = None
                logger.info(f"[port = {self.port}]Connection to Modbus device closed.\n")
            except Exception as e:
                logger.error(f"[port = {self.port}]Error during teardown: {e}\n")
                

def main():
    client = ClientTest()


if __name__ == "__main__":
    main()
